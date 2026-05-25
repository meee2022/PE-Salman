#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Parse جدول الزيارات الأسبوعي - التربية البدنية.xlsx
Output: scripts/activity-import.json
"""

import json
import os
import re
import sys

import openpyxl
from datetime import datetime, date

# ── Paths ──────────────────────────────────────────────────────────────────
EXCEL_PATH = r"C:\Users\M\Desktop\github for local\salman sports\all files\TEAMS\جدول الزيارات الأسبوعي - التربية البدنية.xlsx"
OUTPUT_PATH = r"C:\Users\M\Desktop\github for local\salman sports\scripts\activity-import.json"

# ── Official supervisors ────────────────────────────────────────────────────
OFFICIAL_SUPERVISORS = [
    {"seq": 1,  "name": "سلمان حسن عبدالله الحازمي"},
    {"seq": 2,  "name": "فيصل عبدالعزيز البدر"},
    {"seq": 3,  "name": "أحمد يوسف غريب أبوالفين"},
    {"seq": 4,  "name": "علي حمد محمد الكبيسي"},
    {"seq": 5,  "name": "أحمد صالح أحمد الرئيسي"},
    {"seq": 6,  "name": "مبارك عبدالعزيز ابراهيم المالك"},
    {"seq": 7,  "name": "جميله الماس سعيد الحمد"},
    {"seq": 8,  "name": "اقبال عبدالله عبدالوهاب  السليمان"},
    {"seq": 9,  "name": "حنان خميس حران العمور الشمري"},
    {"seq": 10, "name": "عائشة حمد محمد الشقيري المهندي"},
    {"seq": 11, "name": "هيام صلاح عبدالقادر الغرابات"},
    {"seq": 12, "name": "لينا سعيد محمد زوكاري"},
    {"seq": 13, "name": "موزة عبدالله على محمد الحرمي"},
    {"seq": 14, "name": "فجر محمد صباح الكبيسي"},
]

# ── Name normalization (matches Convex normName) ───────────────────────────
def norm_name(s: str) -> str:
    s = str(s)
    s = s.replace('\xa0', ' ')
    s = s.replace('ـ', '')
    # Remove tashkeel
    s = re.sub(r'[ً-ْٰ]', '', s)
    # Normalize alef variants
    s = re.sub(r'[أإآ]', 'ا', s)
    s = s.replace('ى', 'ي')
    s = s.replace('ة', 'ه')
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

def short_key(s: str) -> str:
    n = norm_name(s)
    parts = n.split(' ')
    if len(parts) >= 2:
        return f"{parts[0]} {parts[-1]}"
    return n

# Build lookup: normalized shortKey → seq
sup_by_short = {}
sup_by_norm = {}
for sup in OFFICIAL_SUPERVISORS:
    sup_by_short[short_key(sup['name'])] = sup['seq']
    sup_by_norm[norm_name(sup['name'])] = sup['seq']

# ── Activity code mapping ──────────────────────────────────────────────────
def classify(cell_text: str):
    """Returns (code, notes_or_None) or None to skip."""
    if cell_text is None:
        return None
    t = str(cell_text).strip()
    # Skip empties / pandas artifacts
    if t in ('', 'nan', 'True', 'False', 'None'):
        return None
    if not t:
        return None

    # Check in order
    if 'مكتب' in t:
        return ('OF', None)
    if 'زيارة' in t:
        return ('VS', t)
    if any(kw in t for kw in ['تطوير', 'ورشة', 'برنامج', 'دورة', 'ارتقاء']):
        return ('TR', t)
    if 'اجتماع' in t:
        return ('MT', t)
    if any(kw in t for kw in ['عن بعد', 'اون لاين', 'اونلاين']):
        return ('OL', t)
    if any(kw in t for kw in ['مرض', 'طبي', 'مرافق']):
        return ('SL', None)
    if 'حج' in t:
        return ('HC', None)
    if any(kw in t for kw in ['رسمي', 'وطني', 'أضح', 'اضح', 'وقفة', 'ميلاد']):
        return ('SP', None)
    if 'عارض' in t:
        return ('CL', None)
    if any(kw in t for kw in ['دوري', 'دروري', 'سنوي']):
        return ('LV', None)
    if 'تقاعد' in t:
        return ('CA', None)
    if 'غياب' in t:
        return ('AB', None)
    if any(kw in t for kw in ['اذن', 'إذن']):
        return ('WP', None)
    if any(kw in t for kw in ['نشاط', 'مسابقة', 'بطولة', 'نهائي', 'معرض',
                                'مؤتمر', 'مهرجان', 'يوم', 'احتفال', 'لياقة', 'كيدز']):
        return ('AC', t)
    if 'مهمة' in t:
        return ('VP', None)
    if 'عاطف' in t:
        return ('CL', None)
    # Anything else with real text → treat as school visit
    return ('VS', t)

# ── Date parsing ───────────────────────────────────────────────────────────
def parse_date_cell(cell_val):
    """
    Extract a date from a header cell.
    Could be a datetime object, a string like '31/8' or '31/8/2025', or an Excel serial.
    Returns a date object or None.
    """
    if cell_val is None:
        return None
    if isinstance(cell_val, (datetime, date)):
        if isinstance(cell_val, datetime):
            return cell_val.date()
        return cell_val

    s = str(cell_val).strip()
    if not s or s in ('nan', 'None', ''):
        return None

    # Try various formats
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    # Try partial 'day/month' → assume 2025 (academic year start) or 2026
    m = re.match(r'^(\d{1,2})[/\-](\d{1,2})$', s)
    if m:
        day, month = int(m.group(1)), int(m.group(2))
        # Academic year 2025-2026: months 8-12 → 2025, months 1-7 → 2026
        year = 2025 if month >= 8 else 2026
        try:
            return date(year, month, day)
        except ValueError:
            pass

    # Excel serial number
    try:
        serial = int(float(s))
        # Excel epoch: 1900-01-01 = 1 (with bug: 1900 treated as leap year, so subtract 2 for serials > 59)
        from datetime import timedelta
        base = date(1899, 12, 30)
        return base + timedelta(days=serial)
    except (ValueError, TypeError, OverflowError):
        pass

    return None

# ── Main parsing logic ─────────────────────────────────────────────────────
def parse_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Sheets found: {wb.sheetnames}", file=sys.stderr)

    logs = []
    skipped_names = set()
    total_cells = 0
    mapped_cells = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 5:
            print(f"  Sheet '{sheet_name}': too few rows ({len(rows)}), skipping", file=sys.stderr)
            continue

        # Row index 3 (0-based) = header row with dates (columns 2-6)
        # Row index 0-2: title rows
        # Row index 4-17: supervisor rows (14 supervisors)
        date_row = rows[3] if len(rows) > 3 else []

        # Extract dates from columns 2-6 (0-indexed)
        # Each column header cell may contain the date
        day_dates = []
        for col_idx in range(2, 7):
            cell_val = date_row[col_idx] if col_idx < len(date_row) else None
            d = parse_date_cell(cell_val)
            day_dates.append(d)

        # If all dates are None, try row 2 or row 1
        if all(d is None for d in day_dates):
            for fallback_row_idx in [2, 1, 0]:
                if fallback_row_idx < len(rows):
                    fallback_row = rows[fallback_row_idx]
                    day_dates_try = []
                    for col_idx in range(2, 7):
                        cell_val = fallback_row[col_idx] if col_idx < len(fallback_row) else None
                        d = parse_date_cell(cell_val)
                        day_dates_try.append(d)
                    if any(d is not None for d in day_dates_try):
                        day_dates = day_dates_try
                        break

        # Supervisor rows: rows 4 to 17 (index 4-17, 14 rows)
        sup_rows = rows[4:18]

        for row in sup_rows:
            if len(row) < 2:
                continue
            name_cell = row[1]
            if name_cell is None:
                continue
            name_str = str(name_cell).strip()
            if not name_str or name_str in ('nan', 'None', ''):
                continue

            # Match supervisor
            norm = norm_name(name_str)
            sk = short_key(name_str)

            seq = None
            if sk in sup_by_short:
                seq = sup_by_short[sk]
            elif norm in sup_by_norm:
                seq = sup_by_norm[norm]
            else:
                # Fuzzy: find best overlap
                best_score = 0
                best_seq = None
                name_words = set(norm.split())
                for sup in OFFICIAL_SUPERVISORS:
                    sup_words = set(norm_name(sup['name']).split())
                    score = len(name_words & sup_words)
                    if score > best_score:
                        best_score = score
                        best_seq = sup['seq']
                threshold = max(1, len(name_words) // 2)
                if best_score >= threshold and best_seq is not None:
                    seq = best_seq

            if seq is None:
                skipped_names.add(name_str)
                continue

            # Process each day column (cols 2-6 → indices 2,3,4,5,6)
            for day_offset, col_idx in enumerate(range(2, 7)):
                if col_idx >= len(row):
                    continue
                cell_val = row[col_idx]
                if cell_val is None:
                    continue
                cell_text = str(cell_val).strip()
                if not cell_text or cell_text in ('nan', 'None', 'True', 'False', ''):
                    continue

                total_cells += 1
                result = classify(cell_text)
                if result is None:
                    continue

                code, notes = result

                # Get date for this column
                d = day_dates[day_offset] if day_offset < len(day_dates) else None
                if d is None:
                    # Can't log without a date
                    print(f"  WARNING: no date for sheet='{sheet_name}' col={col_idx} sup_seq={seq} text='{cell_text}'", file=sys.stderr)
                    continue

                mapped_cells += 1
                entry = {
                    "supervisorSeq": seq,
                    "date": d.isoformat(),
                    "code": code,
                }
                if notes is not None:
                    entry["notes"] = notes
                logs.append(entry)

    print(f"\nTotal non-empty cells processed: {total_cells}", file=sys.stderr)
    print(f"Mapped to codes: {mapped_cells}", file=sys.stderr)
    print(f"Skipped supervisor names: {skipped_names}", file=sys.stderr)

    # Remove duplicate (supervisorSeq, date) — keep last
    seen = {}
    for log in logs:
        key = (log['supervisorSeq'], log['date'])
        seen[key] = log
    deduped = list(seen.values())
    deduped.sort(key=lambda x: (x['supervisorSeq'], x['date']))

    print(f"Unique logs after dedup: {len(deduped)}", file=sys.stderr)

    output = {
        "academicYear": "2025-2026",
        "supervisors": OFFICIAL_SUPERVISORS,
        "logs": deduped,
    }

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\nWrote {len(deduped)} logs to: {OUTPUT_PATH}", file=sys.stderr)
    return deduped

if __name__ == '__main__':
    parse_excel()
